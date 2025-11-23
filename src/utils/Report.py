# PERFORMANCE ANALYSIS REPORT
import pandas as pd
import json
import os
import traceback
import plotly.graph_objects as go
import plotly.express as px
import matplotlib.pyplot as plt
import matplotlib as mpl
from matplotlib.patches import FancyBboxPatch
import matplotlib.ticker as mtick
import textwrap
from plotly.subplots import make_subplots
from IPython.display import display, HTML
from experiment_manager import ExperimentManager
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, List, Optional
import numpy as np

# Patch for expmanager, we load config details and handle the structures
@staticmethod
def load_all_experiments_corrected(base_dir: str = "experiments") -> pd.DataFrame:
    all_metadata = []
    exp_dirs = ExperimentManager._get_all_exp_dirs(base_dir)
    class_names = ['Infected', 'Healthy', 'WBC'] # Define class names once

    for exp_dir in exp_dirs:
        metadata_path = os.path.join(exp_dir, 'metadata.json')
        config_path = os.path.join(exp_dir, 'config.json')
        if os.path.exists(metadata_path) and os.path.exists(config_path):
            try:
                with open(metadata_path, 'r') as f: metadata = json.load(f)
                with open(config_path, 'r') as f: config = json.load(f)

                metadata['directory'] = exp_dir
                metadata['num_classes_detection'] = config.get('num_classes_detection')
                metadata['num_classes_classif'] = config.get('num_classes_classif')
                metadata['num_classes_severity'] = len(config.get('severity_thresholds', [])) + 1 if 'severity' in config.get('active_tasks', []) else None
                
                results_path = None
                mttl_results_path = os.path.join(exp_dir, 'mttl_test_results.json')
                stl_results_path = os.path.join(exp_dir, 'test_results.json')
                if os.path.exists(mttl_results_path): results_path = mttl_results_path
                elif os.path.exists(stl_results_path): results_path = stl_results_path

                if results_path:
                    with open(results_path, 'r') as f:
                        results = json.load(f)

                    def flatten_results_recursive(data, prefix=""):
                        flat_dict = {}
                        for key, value in data.items():
                            new_prefix = f"{prefix}{key}_"
                            if isinstance(value, dict):
                                flat_dict.update(flatten_results_recursive(value, new_prefix))
                            elif isinstance(value, list) and ("_per_class" in key):
                                for i, item in enumerate(value):
                                    if i < len(class_names):
                                        flat_dict[f"{prefix}{key}_{class_names[i]}"] = item
                            else:
                                flat_dict[f"{prefix}{key}"] = value
                        return flat_dict

                    if metadata.get('mode') == 'MTTL':
                        metadata.update(flatten_results_recursive(results))
                    else: 
                        task_name = metadata.get('tasks', ['unknown'])[0]
                        metadata.update(flatten_results_recursive(results, prefix=f"{task_name}_"))
                
                all_metadata.append(metadata)
            except Exception as e:
                print(f"Warning: Could not read metadata/results in {exp_dir}. Error: {e}")

    if not all_metadata: return pd.DataFrame()
    df = pd.DataFrame(all_metadata)
    if 'timestamp_start' in df.columns:
        df['timestamp_start'] = pd.to_datetime(df['timestamp_start'])
        df = df.sort_values(by='timestamp_start', ascending=False).reset_index(drop=True)
    return df

# Apply the patch
ExperimentManager.load_all_experiments = load_all_experiments_corrected

# Report Generator 
class ReportGenerator:
    def __init__(self, base_dir="experiments"):
        print("Initializing Report Generator...")
        self.class_names = ['Infected', 'Healthy', 'WBC']
        raw_df = ExperimentManager.load_all_experiments(base_dir)
        self.all_runs_df = self._preprocess_data(raw_df)
        if self.all_runs_df.empty:
            print("Warning: No completed experiment records found after preprocessing.")
        else:
            print(f"Loaded and preprocessed {len(self.all_runs_df)} completed experiment records.")

    def _preprocess_data(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty or 'status' not in df.columns: return pd.DataFrame()
        df = df[df['status'] == 'completed'].copy()
        if df.empty: return df
        df['Run Name'] = df['strategy']
        
        # Unique name for plotting
        def create_unique_plot_name(row):
            mode = row.get('mode', 'EXP')
            run_id = row.get('run_id', 'X')
            strategy = row.get('strategy', 'Unknown Strategy')
            return f"[{mode} Run {run_id}] {strategy}"
        
        df['unique_plot_name'] = df.apply(create_unique_plot_name, axis=1)
        
        # We combine metrics smartly
        cond1 = df['num_classes_detection'].fillna(-1) == 1
        
        def has_detection_task(task_list):
            if isinstance(task_list, list):
                return 'detection' in task_list
            return False
            
        cond2 = df['tasks'].apply(has_detection_task)
        
        # Create mask
        single_class_det_mask = cond1 & cond2
        
        metric_map = {
            'detection_optimal_f1_score': 'detection_f1_per_class_Infected',
            'detection_optimal_recall': 'detection_recall_per_class_Infected',
            'detection_optimal_precision': 'detection_precision_per_class_Infected'
        }
        
        for source_col, target_col in metric_map.items():
            if source_col in df.columns:
                df.loc[single_class_det_mask, target_col] = df.loc[single_class_det_mask, source_col]

        return df

    @staticmethod
    def _style_table(df: pd.DataFrame, title: str, rank_metric: str, col_map: dict):
        print(f"\n{'='*110}\n{title.upper()}\n{'='*110}")
        
        df_display_shell = pd.DataFrame(columns=col_map.values())

        if df.empty:
            print("No data available for this report.")
            display(df_display_shell); return

        df = df.copy()

        if 'tasks' in df.columns:
            df['tasks'] = df['tasks'].apply(lambda x: ', '.join(x) if isinstance(x, list) else str(x))

        existing_cols = [col for col in col_map.keys() if col in df.columns]
        missing_cols = [col_map[col] for col in col_map.keys() if col not in df.columns]

        if not existing_cols:
            print("Cannot generate report: None of the required columns were found in the data.")
            display(df_display_shell); return
        
        if missing_cols:
            print(f"Note: Some columns could not be found in results and will be omitted: {missing_cols}")

        df_display = df[existing_cols].copy().rename(columns=col_map)
        
        rank_metric_renamed = col_map.get(rank_metric)
        if rank_metric_renamed and rank_metric_renamed in df_display.columns and df_display[rank_metric_renamed].notna().any():
            print(f"(Ranked by Highest: {rank_metric_renamed})\n")
            df_display = df_display.sort_values(by=rank_metric_renamed, ascending=False)
        else:
            print(f"(Unsorted - Ranking metric '{rank_metric_renamed}' not found or has no data)\n")

        df_display.index = np.arange(1, len(df_display) + 1)
        df_display.index.name = "Rank"

        for col in df_display.columns:
            if df_display[col].dtype in ['float64', 'float32', 'float16']:
                df_display[col] = df_display[col].apply(lambda x: f"{x:.4f}" if pd.notna(x) else '-')
            if col == 'Dir':
                df_display[col] = df_display[col].apply(lambda x: os.path.join("...", *x.split(os.sep)[-3:]) if isinstance(x, str) else x)
        
        display(df_display.style.set_properties(**{'text-align': 'left', 'white-space': 'nowrap'}).format(na_rep='-'))
    
    def _get_table_configs(self):
        """Central definition for all table structures."""
        return {
            'L-Det': ('detection_f1_per_class_Infected', {
                'Run Name': 'Run', 'tasks': 'Tasks', 'num_classes_detection': '#Cls',
                'detection_mAP_50': 'mAP@50', 'detection_mAP_75': 'mAP@75',
                'detection_f1_per_class_Infected': 'F1(Inf)', 'detection_recall_per_class_Infected': 'R(Inf)', 'detection_precision_per_class_Infected': 'P(Inf)',
                'directory': 'Dir'
            }),
            'L-Heatmap': ('heatmap_dice', {
                'Run Name': 'Run', 'tasks': 'Tasks', 'heatmap_dice': 'dice',
                'heatmap_mse': 'mse', 'heatmap_mae': 'mae', 'heatmap_correlation': 'correlation', 'directory': 'Dir'
            }),
            'L-Classif': ('f1_macro', {
                'Run Name': 'Run', 'tasks': 'Tasks', 'num_classes_classif': '#Cls',
                'accuracy': 'Acc', 'f1_macro': 'F1-Macro', 'directory': 'Dir'
            }),
            'L-Seg': ('segmentation_dice', {
                'Run Name': 'Run', 'tasks': 'Tasks', 'segmentation_dice': 'Dice',
                'segmentation_iou': 'IoU', 'directory': 'Dir'
            }),
            'R-DetSingle': ('detection_optimal_f1_score', {
                'Run Name': 'Run', 'tasks': 'Tasks', 'num_classes_detection': '#Cls',
                'detection_mAP_50': 'mAP@50', 'detection_mAP_75': 'mAP@75', 'detection_optimal_confidence_threshold': 'Conf(Opt)',
                'detection_optimal_f1_score': 'F1(Opt)', 'detection_optimal_recall': 'R(Opt)', 'detection_optimal_precision': 'P(Opt)',
                'detection_count_accuracy_percent': 'CountAcc(Inf)%', 'directory': 'Dir'
            }),
            'R-DetMulti': ('detection_f1_per_class_Infected', {
                'Run Name': 'Run', 'tasks': 'Tasks', 'num_classes_detection': '#Cls',
                'detection_mAP_50': 'mAP@50', 'detection_mAP_75': 'mAP@75',
                'detection_optimal_confidence_threshold': 'Conf(Opt)',  
                'detection_optimal_f1_score': 'F1(Opt)',              
                'detection_optimal_recall': 'R(Opt)',                
                'detection_optimal_precision': 'P(Opt)',             
                'detection_f1_per_class_Infected': 'F1(Inf)', 'detection_f1_per_class_Healthy': 'F1(H)',
                'detection_recall_per_class_Infected': 'R(Inf)', 'detection_recall_per_class_Healthy': 'R(H)',
                'detection_precision_per_class_Infected': 'P(Inf)', 'detection_precision_per_class_Healthy': 'P(H)',
                'detection_count_accuracy_per_class_percent_Infected': 'CountAcc(Inf)%', 'directory': 'Dir'
            }),
        }

    def generate_all_reports(self):
        if self.all_runs_df.empty:
            print("Analysis skipped: No completed experiments found.")
            return
        
        self._generate_leaderboards()
        self._display_full_report_for_mode('STL')
        self._display_full_report_for_mode('MTTL')
        self.plot_top_performers()

    def _filter_df_for_task(self, df, task_name):
        if df.empty: return pd.DataFrame()
        return df[df['tasks'].apply(lambda tasks: isinstance(tasks, list) and task_name in tasks)].copy()

    def _generate_leaderboards(self):
        print(f"\n\n{'*'*45} SOTA LEADERBOARDS (ALL MODES) {'*'*45}")
        df = self.all_runs_df
        configs = self._get_table_configs()
        
        self._style_table(self._filter_df_for_task(df, 'detection'), "Leaderboard: Detection", *configs['L-Det'])
        self._style_table(self._filter_df_for_task(df, 'segmentation'), "Leaderboard: Segmentation", *configs['L-Seg'])
        self._style_table(self._filter_df_for_task(df, 'heatmap'), "Leaderboard: Heatmap Localization", *configs['L-Heatmap'])
        
        for prefix, name in [('roi_classif', 'ROI'), ('cell_classif', 'Cell'), ('severity', 'Severity')]:
            df_task = self._filter_df_for_task(df, prefix)
            if not df_task.empty:
                df_task_renamed = df_task.copy().rename(columns=lambda c: c.replace(f'{prefix}_', ''))
                if 'severity' in prefix: df_task_renamed.rename(columns={'num_classes_severity': 'num_classes_classif'}, inplace=True)
                self._style_table(df_task_renamed, f"Leaderboard: {name} Classification", *configs['L-Classif'])

    def _display_full_report_for_mode(self, mode: str):
        print(f"\n\n{'='*40} {mode} MODE: INDIVIDUAL TASK REPORTS {'='*40}")
        mode_df = self.all_runs_df[self.all_runs_df['mode'] == mode]
        if mode_df.empty: print(f"No completed experiments found for {mode} mode."); return
        configs = self._get_table_configs()
        
        det_df = self._filter_df_for_task(mode_df, 'detection')
        self._style_table(det_df[det_df['num_classes_detection'] == 1], f"{mode} Report: Detection (Single-Class)", *configs['R-DetSingle'])
        self._style_table(det_df[det_df['num_classes_detection'] > 1], f"{mode} Report: Detection (Multi-Class)", *configs['R-DetMulti'])
        
        self._style_table(self._filter_df_for_task(mode_df, 'segmentation'), f"{mode} Report: Segmentation", *configs['L-Seg'])
        self._style_table(self._filter_df_for_task(mode_df, 'heatmap'), f"{mode} Report: Heatmap Localization", *configs['L-Heatmap'])

        for prefix, name in [('roi_classif', 'ROI'), ('cell_classif', 'Cell'), ('severity', 'Severity')]:
            df_task = self._filter_df_for_task(mode_df, prefix)
            if not df_task.empty:
                df_task_renamed = df_task.copy().rename(columns=lambda c: c.replace(f'{prefix}_', ''))
                if 'severity' in prefix: df_task_renamed.rename(columns={'num_classes_severity': 'num_classes_classif'}, inplace=True)
                self._style_table(df_task_renamed, f"{mode} Report: {name} Classification", *configs['L-Classif'])

    def plot_top_performers(self):
        print(f"\n\n{'='*45} TOP PERFORMERS VISUALIZATION (ADVANCED) {'='*45}")
        df = self.all_runs_df
        if df.empty:
            print("No data to visualize.")
            return

        # Task and metrics
        task_configs = {
            'Detection': {
                'task_key': 'detection',
                'rank_metric': 'detection_f1_per_class_Infected',
                'metrics': {
                    'mAP@50': 'detection_mAP_50',
                    'F1 (Inf)': 'detection_f1_per_class_Infected',
                    'Recall (Inf)': 'detection_recall_per_class_Infected',
                    'Precision (Inf)': 'detection_precision_per_class_Infected'
                }
            },
            'Segmentation': {
                'task_key': 'segmentation',
                'rank_metric': 'segmentation_dice',
                'metrics': {
                    'Dice': 'segmentation_dice',
                    'IoU': 'segmentation_iou'
                }
            },
            'Heatmap Localization': {
                'task_key': 'heatmap',
                'rank_metric': 'heatmap_dice',
                'metrics': {
                    'Dice': 'heatmap_dice',
                    'Correlation': 'heatmap_correlation',
                    '1 - MAE': 'heatmap_mae' # Inverted
                }
            },
            'ROI Classification': {
                'task_key': 'roi_classif',
                'rank_metric': 'roi_classif_f1_macro',
                'metrics': {
                    'F1-Macro': 'roi_classif_f1_macro',
                    'Accuracy': 'roi_classif_accuracy'
                }
            }
        }
        
        if 'heatmap_mae' in df.columns:
            df['heatmap_mae'] = 1 - df['heatmap_mae']

        # plot
        for task_title, config in task_configs.items():
            df_task = self._filter_df_for_task(df, config['task_key'])
            
            # Check if there's any data to plot
            if df_task.empty or not any(m in df_task.columns for m in config['metrics'].values()):
                print(f"\nSkipping '{task_title}' plots: No relevant data found.")
                continue

            top_5 = df_task.sort_values(by=config['rank_metric'], ascending=False).head(5)
            
            # Prepare data for plotting
            metric_labels = list(config['metrics'].keys())
            metric_cols = list(config['metrics'].values())
            plot_data = top_5[['unique_plot_name'] + metric_cols].copy()
            
            # Clean unique_plot_names for better display
            plot_data['unique_plot_name'] = plot_data['unique_plot_name'].str.replace('_', ' ').str.wrap(25)

            # Multi-Metric Grouped Bar Chart
            fig_bar = go.Figure()
            colors = px.colors.qualitative.Plotly
            
            for i, metric_label in enumerate(metric_labels):
                metric_col = config['metrics'][metric_label]
                if metric_col in plot_data.columns:
                    fig_bar.add_trace(go.Bar(
                        name=metric_label,
                        x=plot_data['unique_plot_name'],
                        y=plot_data[metric_col],
                        text=[f"{v:.3f}" for v in plot_data[metric_col]],
                        textposition='outside',
                        marker_color=colors[i % len(colors)]
                    ))

            fig_bar.update_layout(
                title=f'<b>{task_title}: Top 5 Performers (Multi-Metric Comparison)</b>',
                xaxis_tickfont_size=12,
                yaxis=dict(title='Score', range=[0, 1.1], tickformat=".2f"),
                barmode='group',
                legend_title_text='Metrics',
                template='plotly_white',
                height=600,
                uniformtext_minsize=8,
                uniformtext_mode='hide'
            )
            fig_bar.show()

            # Multi-Metric Radar Chart
            fig_radar = go.Figure()
            
            for i in range(len(top_5)):
                run_data = top_5.iloc[i]
                values = [run_data.get(col, 0) for col in metric_cols] # 0 if missing
                
                fig_radar.add_trace(go.Scatterpolar(
                    r=values,
                    theta=metric_labels,
                    fill='toself',
                    name=textwrap.shorten(run_data['unique_plot_name'], width=40, placeholder="…"),
                    hovertemplate='<b>%{customdata[0]}</b><br>%{theta}: %{r:.4f}<extra></extra>',
                    customdata=[[run_data['unique_plot_name']]]
                ))

            fig_radar.update_layout(
                title=f'<b>{task_title}: Performance Fingerprint (Top 5 Models)</b>',
                polar=dict(
                    radialaxis=dict(visible=True, range=[0, 1])
                ),
                legend_title_text='Models',
                template='plotly_white',
                height=700
            )
            fig_radar.show()

class ColorScheme:
    """ color palette."""
    HEADER_DARK = '1F4E78'
    HEADER_LIGHT = 'D9E1F2'
    ACCENT_BLUE = '4472C4'
    ACCENT_GREEN = '70AD47'
    ACCENT_RED = 'C55A11'
    GOLD = 'FFC000'
    SILVER = 'C0C0C0'
    BRONZE = 'CD7F32'
    LIGHT_GRAY = 'F2F2F2'
    WHITE = 'FFFFFF'
    GAIN_POS = 'C6EFCE'  # Light green for positive gains
    GAIN_NEG = 'FFC7CE'  # Light red for negative gains

class FontManager:
    """Centralized font definitions."""
    def __init__(self):
        self.title = Font(name='Calibri', size=18, bold=True, color='FFFFFF')
        self.section = Font(name='Calibri', size=13, bold=True, color='1F4E78')
        self.header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        self.subheader = Font(name='Calibri', size=10, bold=True, color='1F4E78')
        self.regular = Font(name='Calibri', size=10)
        self.bold = Font(name='Calibri', size=10, bold=True)
        self.small = Font(name='Calibri', size=9)

class BorderManager:
    """Border definitions."""
    THIN = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    MEDIUM = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )

class ExcelSheetBuilder:    
    def __init__(self, workbook: openpyxl.Workbook, sheet_name: str):
        self.ws = workbook.create_sheet(sheet_name)
        self.colors = ColorScheme()
        self.fonts = FontManager()
        self.border = BorderManager.THIN
        self.current_row = 1
        
    def add_title(self, text: str, col_start: str = 'A', col_end: str = 'H') -> int:
        """Add a title with background color."""
        cell = self.ws[f'{col_start}{self.current_row}']
        cell.value = text
        cell.font = self.fonts.title
        cell.fill = PatternFill(start_color=self.colors.ACCENT_BLUE, 
                               end_color=self.colors.ACCENT_BLUE, fill_type='solid')
        self.ws.merge_cells(f'{col_start}{self.current_row}:{col_end}{self.current_row}')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        self.ws.row_dimensions[self.current_row].height = 28
        self.current_row += 2
        return self.current_row
    
    def add_section(self, text: str, col_end: str = 'H') -> int:
        """Add a section header."""
        cell = self.ws[f'A{self.current_row}']
        cell.value = text
        cell.font = self.fonts.section
        cell.fill = PatternFill(start_color=self.colors.HEADER_LIGHT, 
                               end_color=self.colors.HEADER_LIGHT, fill_type='solid')
        self.ws.merge_cells(f'A{self.current_row}:{col_end}{self.current_row}')
        cell.border = self.border
        self.ws.row_dimensions[self.current_row].height = 22
        self.current_row += 1
        return self.current_row
    
    def add_header_row(self, columns: List[str], start_col: int = 1) -> int:
        """Add a header row with  formatting."""
        row = self.current_row
        for col_idx, col_name in enumerate(columns, start_col):
            cell = self.ws.cell(row=row, column=col_idx)
            cell.value = col_name.replace('_', ' ').title()
            cell.font = self.fonts.header
            cell.fill = PatternFill(start_color=self.colors.HEADER_DARK, 
                                   end_color=self.colors.HEADER_DARK, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
        
        self.ws.row_dimensions[row].height = 24
        self.current_row += 1
        return self.current_row
    
    def add_data_row(self, values: List, rank: Optional[int] = None, highlight: bool = False, 
                     gain_color: Optional[str] = None, start_col: int = 1) -> None:
        """Add a data row with conditional formatting."""
        row = self.current_row
        for col_idx, value in enumerate(values, start_col):
            cell = self.ws.cell(row=row, column=col_idx)
            
            if isinstance(value, float):
                if value > 100 or value < -100:  # Likely a percentage gain
                    cell.value = round(value, 2)
                    cell.number_format = '0.00'
                else:
                    cell.value = round(value, 4)
                    cell.number_format = '0.0000'
            elif isinstance(value, list):
                cell.value = ', '.join(str(v) for v in value)
            else:
                cell.value = value
            
            cell.font = self.fonts.regular
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Rank-based coloring
            if rank == 1:
                cell.fill = PatternFill(start_color=self.colors.GOLD, 
                                       end_color=self.colors.GOLD, fill_type='solid')
                cell.font = self.fonts.bold
            elif rank == 2:
                cell.fill = PatternFill(start_color=self.colors.SILVER, 
                                       end_color=self.colors.SILVER, fill_type='solid')
                cell.font = self.fonts.bold
            elif rank == 3:
                cell.fill = PatternFill(start_color=self.colors.BRONZE, 
                                       end_color=self.colors.BRONZE, fill_type='solid')
                cell.font = self.fonts.bold
            elif gain_color:
                cell.fill = PatternFill(start_color=gain_color, 
                                       end_color=gain_color, fill_type='solid')
            elif highlight:
                cell.fill = PatternFill(start_color=self.colors.HEADER_LIGHT, 
                                       end_color=self.colors.HEADER_LIGHT, fill_type='solid')
            elif row % 2 == 0:
                cell.fill = PatternFill(start_color=self.colors.LIGHT_GRAY, 
                                       end_color=self.colors.LIGHT_GRAY, fill_type='solid')
        
        self.ws.row_dimensions[row].height = 18
        self.current_row += 1
    
    def add_table_with_filtering(self, data: pd.DataFrame, columns: List[str], 
                                  start_col: int = 1, include_rank: bool = False) -> None:
        """Add a table with built-in filtering capability."""
        if data.empty:
            return
        
        # Prepare data
        df_display = data[columns].copy()
        df_display.columns = [c.replace('_', ' ').title() for c in columns]
        
        # Add rank column if requested
        if include_rank:
            df_display.insert(0, 'Rank', range(1, len(df_display) + 1))
        
        # Add headers
        row = self.current_row
        header_cols = ['Rank'] + list(df_display.columns) if include_rank else list(df_display.columns)
        self.add_header_row(header_cols, start_col=start_col)
        
        # Add data rows
        for idx, (_, row_data) in enumerate(df_display.iterrows(), 1):
            values = [idx] + row_data.tolist() if include_rank else row_data.tolist()
            self.add_data_row(values, start_col=start_col)
        
        # Create Excel table with filtering
        end_col = get_column_letter(start_col + len(df_display.columns) + (1 if include_rank else 0) - 1)
        table_ref = f'{get_column_letter(start_col)}{row}:{end_col}{self.current_row - 1}'
        
        # Sanitize table name
        sanitized_name = self.ws.title.replace(' ', '_').replace('-', '_')
        table_display_name = f'Table_{sanitized_name}_{row}'
        
        try:
            tab = Table(displayName=table_display_name, ref=table_ref)
            style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab.tableStyleInfo = style
            self.ws.add_table(tab)
        except:
            pass  # Table creation might fail on some versions
        
        self.current_row += 1
    
    def auto_adjust_columns(self) -> None:
        """Auto-adjust column widths based on content."""
        for col_idx in range(1, self.ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for row in self.ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if isinstance(cell, openpyxl.cell.cell.MergedCell):
                        continue
                    try:
                        cell_value = str(cell.value) if cell.value else ""
                        max_length = max(max_length, len(cell_value))
                    except:
                        pass
            
            adjusted_width = min(max_length + 3, 60)
            self.ws.column_dimensions[column_letter].width = adjusted_width

class SummarySheetBuilder(ExcelSheetBuilder):
    """Executive Summary sheet."""
    
    def __init__(self, workbook: openpyxl.Workbook):
        super().__init__(workbook, "Summary")
    
    def build(self, report_gen) -> None:
        """Build the summary sheet."""
        self.add_title("PERFORMANCE ANALYSIS: EXECUTIVE SUMMARY")
        
        # Metadata section
        self.add_section("Report Metadata")
        self._add_metadata(report_gen)
        
        # Mode summary
        self.current_row += 1
        self.add_section("Experiment Distribution by Mode")
        self._add_mode_summary(report_gen)
        
        # Task distribution
        self.current_row += 1
        self.add_section("Task Distribution")
        self._add_task_distribution(report_gen)
        
        # Best performers
        self.current_row += 1
        self.add_section("Best Performers by Task")
        self._add_best_performers(report_gen)
        
        self.auto_adjust_columns()
    
    def _add_metadata(self, report_gen) -> None:
        """Add metadata rows."""
        metadata = [
            ("Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Total Experiments", len(report_gen.all_runs_df)),
            ("Completed", len(report_gen.all_runs_df[report_gen.all_runs_df['status'] == 'completed'])),
        ]
        
        for label, value in metadata:
            row = self.current_row
            self.ws[f'A{row}'] = label
            self.ws[f'B{row}'] = value
            self.ws[f'A{row}'].font = self.fonts.bold
            self.ws[f'A{row}'].fill = PatternFill(start_color=self.colors.LIGHT_GRAY, 
                                                  end_color=self.colors.LIGHT_GRAY, fill_type='solid')
            self.ws[f'B{row}'].border = self.border
            self.current_row += 1
    
    def _add_mode_summary(self, report_gen) -> None:
        """Add mode distribution."""
        columns = ['Mode', 'Count', 'Percentage']
        self.add_header_row(columns)
        
        mode_counts = report_gen.all_runs_df['mode'].value_counts()
        total = len(report_gen.all_runs_df)
        
        for mode in sorted(mode_counts.index):
            count = mode_counts[mode]
            pct = count / total
            values = [mode, count, pct]
            self.add_data_row(values, highlight=True)
    
    def _add_task_distribution(self, report_gen) -> None:
        """Add task distribution."""
        columns = ['Task', 'Count', 'Percentage']
        self.add_header_row(columns)
        
        task_counts = {}
        for tasks_list in report_gen.all_runs_df['tasks']:
            if isinstance(tasks_list, list):
                for task in tasks_list:
                    task_counts[task] = task_counts.get(task, 0) + 1
        
        total = len(report_gen.all_runs_df)
        for task in sorted(task_counts.keys()):
            count = task_counts[task]
            pct = count / total
            values = [task.replace('_', ' ').title(), count, pct]
            self.add_data_row(values)
    
    def _add_best_performers(self, report_gen) -> None:
        """Add best performers by task."""
        columns = ['Task', 'Best Strategy', 'Mode', 'Key Metric', 'Score']
        self.add_header_row(columns)
        
        tasks_to_check = ['detection', 'segmentation', 'heatmap', 'roi_classif']
        rank_metrics = {
            'detection': 'detection_f1_per_class_Infected',
            'segmentation': 'segmentation_dice',
            'heatmap': 'heatmap_dice',
            'roi_classif': 'roi_classif_f1_macro'
        }
        
        for task in tasks_to_check:
            df_task = report_gen._filter_df_for_task(report_gen.all_runs_df, task)
            if df_task.empty:
                continue
            
            rank_metric = rank_metrics.get(task)
            if rank_metric not in df_task.columns:
                continue
            
            best = df_task.nlargest(1, rank_metric).iloc[0]
            values = [
                task.replace('_', ' ').title(),
                best['strategy'],
                best['mode'],
                rank_metric.replace('_', ' ').title(),
                best[rank_metric]
            ]
            self.add_data_row(values, rank=1)

class LeaderboardSheetBuilder(ExcelSheetBuilder):
    """leaderboard sheet with all tasks."""
    
    METRIC_COLUMNS = {
        'Detection_1cls': [
            'strategy', 'mode', 'detection_mAP_50', 'detection_mAP_75',
            'detection_f1_per_class_Infected', 'detection_recall_per_class_Infected',
            'detection_precision_per_class_Infected'
        ],
        'Detection_2cls': [
            'strategy', 'mode', 'detection_mAP_50', 'detection_mAP_75',
            'detection_f1_per_class_Infected', 'detection_f1_per_class_Healthy',
            'detection_recall_per_class_Infected', 'detection_precision_per_class_Infected'
        ],
        'Detection_3cls': [
            'strategy', 'mode', 'detection_mAP_50', 'detection_mAP_75',
            'detection_f1_per_class_Infected', 'detection_f1_per_class_Healthy',
            'detection_recall_per_class_Infected', 'detection_precision_per_class_Infected'
        ],
        'Segmentation': [
            'strategy', 'mode', 'segmentation_dice', 'segmentation_iou'
        ],
        'Heatmap': [
            'strategy', 'mode', 'heatmap_dice', 'heatmap_correlation', 'heatmap_mae'
        ],
        'ROI_2cls': [
            'strategy', 'mode', 'roi_classif_f1_macro', 'roi_classif_accuracy', 'roi_classif_recall'
        ],
        'ROI_3cls': [
            'strategy', 'mode', 'roi_classif_f1_macro', 'roi_classif_accuracy', 'roi_classif_recall'
        ]
    }
    
    def __init__(self, workbook: openpyxl.Workbook):
        super().__init__(workbook, "Leaderboards")
    
    def build(self, report_gen) -> None:
        """Build unified leaderboards."""
        self.add_title("GLOBAL LEADERBOARDS: ALL TASKS")
        
        # Detection by classes
        for num_cls in [1, 2, 3]:
            det_df = report_gen.all_runs_df[
                (report_gen.all_runs_df['tasks'].apply(lambda t: 'detection' in t if isinstance(t, list) else False)) &
                (report_gen.all_runs_df['num_classes_detection'] == num_cls)
            ].dropna(subset=['detection_f1_per_class_Infected'])
            
            if not det_df.empty:
                self.current_row += 1
                metric_key = f'Detection_{num_cls}cls'
                self.add_section(f"Detection Leaderboard - {num_cls} Class{'es' if num_cls > 1 else ''}")
                self._add_table(det_df, metric_key, 'detection_f1_per_class_Infected')
        
        # Segmentation
        seg_df = report_gen._filter_df_for_task(report_gen.all_runs_df, 'segmentation')
        if not seg_df.empty and 'segmentation_dice' in seg_df.columns:
            self.current_row += 1
            self.add_section("Segmentation Leaderboard")
            self._add_table(seg_df, 'Segmentation', 'segmentation_dice')
        
        # Heatmap
        heat_df = report_gen._filter_df_for_task(report_gen.all_runs_df, 'heatmap')
        if not heat_df.empty and 'heatmap_dice' in heat_df.columns:
            self.current_row += 1
            self.add_section("Heatmap Localization Leaderboard")
            self._add_table(heat_df, 'Heatmap', 'heatmap_dice')
        
        # ROI Classification
        for num_cls in [2, 3]:
            roi_df = report_gen._filter_df_for_task(report_gen.all_runs_df, 'roi_classif')
            roi_cls_df = roi_df.dropna(subset=['num_classes_classif'])
            roi_cls_df = roi_cls_df[roi_cls_df['num_classes_classif'].astype(int) == num_cls]
            
            if not roi_cls_df.empty and 'roi_classif_f1_macro' in roi_cls_df.columns:
                self.current_row += 1
                metric_key = f'ROI_{num_cls}cls'
                self.add_section(f"ROI Classification Leaderboard - {num_cls} Class{'es' if num_cls > 1 else ''}")
                self._add_table(roi_cls_df, metric_key, 'roi_classif_f1_macro')
        
        self.auto_adjust_columns()
    
    def _add_table(self, df: pd.DataFrame, metric_key: str, rank_metric: str) -> None:
        """Add a leaderboard table."""
        columns = self.METRIC_COLUMNS.get(metric_key, ['strategy', 'mode'])
        columns = [c for c in columns if c in df.columns]
        
        df_sorted = df[columns].copy()
        df_sorted = df_sorted.sort_values(by=rank_metric, ascending=False).reset_index(drop=True)
        
        self.add_table_with_filtering(df_sorted, columns, include_rank=True)

class STLDetailsSheetBuilder(ExcelSheetBuilder):
    """STL mode details with filtering."""
    
    METRIC_COLUMNS = {
        'Detection': ['strategy', 'num_classes_detection', 'detection_mAP_50', 'detection_mAP_75',
                     'detection_f1_per_class_Infected', 'detection_recall_per_class_Infected',
                     'detection_precision_per_class_Infected'],
        'Segmentation': ['strategy', 'segmentation_dice', 'segmentation_iou'],
        'Heatmap': ['strategy', 'heatmap_dice', 'heatmap_correlation', 'heatmap_mae'],
        'ROI': ['strategy', 'num_classes_classif', 'roi_classif_f1_macro', 'roi_classif_accuracy']
    }
    
    def __init__(self, workbook: openpyxl.Workbook):
        super().__init__(workbook, "STL Details")
    
    def build(self, report_gen) -> None:
        """Build STL details."""
        stl_df = report_gen.all_runs_df[report_gen.all_runs_df['mode'] == 'STL']
        
        if stl_df.empty:
            return
        
        self.add_title("STL MODE: DETAILED RESULTS (Filterable)")
        
        # Detection
        det_df = report_gen._filter_df_for_task(stl_df, 'detection')
        if not det_df.empty:
            self.current_row += 1
            self.add_section("Detection Results")
            self._add_filtered_table(det_df, 'Detection')
        
        # Segmentation
        seg_df = report_gen._filter_df_for_task(stl_df, 'segmentation')
        if not seg_df.empty:
            self.current_row += 1
            self.add_section("Segmentation Results")
            self._add_filtered_table(seg_df, 'Segmentation')
        
        # Heatmap
        heat_df = report_gen._filter_df_for_task(stl_df, 'heatmap')
        if not heat_df.empty:
            self.current_row += 1
            self.add_section("Heatmap Results")
            self._add_filtered_table(heat_df, 'Heatmap')
        
        # ROI
        roi_df = report_gen._filter_df_for_task(stl_df, 'roi_classif')
        if not roi_df.empty:
            self.current_row += 1
            self.add_section("ROI Classification Results")
            self._add_filtered_table(roi_df, 'ROI')
        
        self.auto_adjust_columns()
    
    def _add_filtered_table(self, df: pd.DataFrame, task_key: str) -> None:
        """Add filterable table."""
        columns = self.METRIC_COLUMNS.get(task_key, ['strategy'])
        columns = [c for c in columns if c in df.columns]
        
        df_display = df[columns].copy()
        self.add_table_with_filtering(df_display, columns, include_rank=True)

class MTTLDetailsSheetBuilder(ExcelSheetBuilder):
    """MTTL mode details with filtering."""
    
    METRIC_COLUMNS = {
        'Detection': ['strategy', 'num_classes_detection', 'detection_mAP_50', 'detection_mAP_75',
                     'detection_f1_per_class_Infected', 'detection_recall_per_class_Infected',
                     'detection_precision_per_class_Infected'],
        'Segmentation': ['strategy', 'segmentation_dice', 'segmentation_iou'],
        'Heatmap': ['strategy', 'heatmap_dice', 'heatmap_correlation', 'heatmap_mae'],
        'ROI': ['strategy', 'num_classes_classif', 'roi_classif_f1_macro', 'roi_classif_accuracy']
    }
    
    def __init__(self, workbook: openpyxl.Workbook):
        super().__init__(workbook, "MTTL Details")
    
    def build(self, report_gen) -> None:
        """Build MTTL details."""
        mttl_df = report_gen.all_runs_df[report_gen.all_runs_df['mode'] == 'MTTL']
        
        if mttl_df.empty:
            return
        
        self.add_title("MTTL MODE: DETAILED RESULTS (Filterable)")
        
        # Detection
        det_df = report_gen._filter_df_for_task(mttl_df, 'detection')
        if not det_df.empty:
            self.current_row += 1
            self.add_section("Detection Results")
            self._add_filtered_table(det_df, 'Detection')
        
        # Segmentation
        seg_df = report_gen._filter_df_for_task(mttl_df, 'segmentation')
        if not seg_df.empty:
            self.current_row += 1
            self.add_section("Segmentation Results")
            self._add_filtered_table(seg_df, 'Segmentation')
        
        # Heatmap
        heat_df = report_gen._filter_df_for_task(mttl_df, 'heatmap')
        if not heat_df.empty:
            self.current_row += 1
            self.add_section("Heatmap Results")
            self._add_filtered_table(heat_df, 'Heatmap')
        
        # ROI
        roi_df = report_gen._filter_df_for_task(mttl_df, 'roi_classif')
        if not roi_df.empty:
            self.current_row += 1
            self.add_section("ROI Classification Results")
            self._add_filtered_table(roi_df, 'ROI')
        
        self.auto_adjust_columns()
    
    def _add_filtered_table(self, df: pd.DataFrame, task_key: str) -> None:
        """Add filterable table."""
        columns = self.METRIC_COLUMNS.get(task_key, ['strategy'])
        columns = [c for c in columns if c in df.columns]
        
        df_display = df[columns].copy()
        self.add_table_with_filtering(df_display, columns, include_rank=True)

class ComparisonSheetBuilder(ExcelSheetBuilder):
    """STL vs MTTL Comparison with Gain metrics - Category-based comparison."""
    
    COMPARISON_CONFIGS = {
        'Detection 1-Class': {
            'task': 'detection',
            'num_classes_key': 'num_classes_detection',
            'num_classes_val': 1,
            'metric': 'detection_f1_per_class_Infected',
            'display_metrics': ['detection_mAP_50', 'detection_f1_per_class_Infected', 
                              'detection_recall_per_class_Infected', 'detection_precision_per_class_Infected']
        },
        'Detection 2-Class': {
            'task': 'detection',
            'num_classes_key': 'num_classes_detection',
            'num_classes_val': 2,
            'metric': 'detection_f1_per_class_Infected',
            'display_metrics': ['detection_mAP_50', 'detection_f1_per_class_Infected', 
                              'detection_f1_per_class_Healthy', 'detection_precision_per_class_Infected']
        },
        'Detection 3-Class': {
            'task': 'detection',
            'num_classes_key': 'num_classes_detection',
            'num_classes_val': 3,
            'metric': 'detection_f1_per_class_Infected',
            'display_metrics': ['detection_mAP_50', 'detection_f1_per_class_Infected', 
                              'detection_f1_per_class_Healthy', 'detection_precision_per_class_Infected']
        },
        'Segmentation': {
            'task': 'segmentation',
            'num_classes_key': None,
            'num_classes_val': None,
            'metric': 'segmentation_dice',
            'display_metrics': ['segmentation_dice', 'segmentation_iou']
        },
        'Heatmap': {
            'task': 'heatmap',
            'num_classes_key': None,
            'num_classes_val': None,
            'metric': 'heatmap_dice',
            'display_metrics': ['heatmap_dice', 'heatmap_correlation', 'heatmap_mae']
        },
        'ROI 2-Class': {
            'task': 'roi_classif',
            'num_classes_key': 'num_classes_classif',
            'num_classes_val': 2,
            'metric': 'roi_classif_f1_macro',
            'display_metrics': ['roi_classif_f1_macro', 'roi_classif_accuracy']
        },
        'ROI 3-Class': {
            'task': 'roi_classif',
            'num_classes_key': 'num_classes_classif',
            'num_classes_val': 3,
            'metric': 'roi_classif_f1_macro',
            'display_metrics': ['roi_classif_f1_macro', 'roi_classif_accuracy']
        }
    }
    
    def __init__(self, workbook: openpyxl.Workbook):
        super().__init__(workbook, "STL vs MTTL")
    
    def build(self, report_gen) -> None:
        """Build comparison sheet with category-based comparisons."""
        self.add_title("STL vs MTTL: COMPARATIVE ANALYSIS WITH GAIN METRICS")
        
        for category_name, config in self.COMPARISON_CONFIGS.items():
            # Filter STL data for this category
            stl_df = report_gen._filter_df_for_task(
                report_gen.all_runs_df[report_gen.all_runs_df['mode'] == 'STL'], 
                config['task']
            )
            
            # Filter MTTL data for this category
            mttl_df = report_gen._filter_df_for_task(
                report_gen.all_runs_df[report_gen.all_runs_df['mode'] == 'MTTL'], 
                config['task']
            )
            
            # Apply class filtering if applicable
            if config['num_classes_key'] is not None and config['num_classes_val'] is not None:
                stl_df = stl_df.dropna(subset=[config['num_classes_key']])
                stl_df = stl_df[stl_df[config['num_classes_key']].astype(int) == config['num_classes_val']]
                
                mttl_df = mttl_df.dropna(subset=[config['num_classes_key']])
                mttl_df = mttl_df[mttl_df[config['num_classes_key']].astype(int) == config['num_classes_val']]
            
            if stl_df.empty or mttl_df.empty:
                continue
            
            self.current_row += 1
            self.add_section(f"{category_name} Comparison")
            self._add_comparison_table(stl_df, mttl_df, config)
        
        self.auto_adjust_columns()
    
    def _add_comparison_table(self, stl_df: pd.DataFrame, mttl_df: pd.DataFrame, config: dict) -> None:
        """Add comparison table with individual metric gains vs best STL baseline with color coding and filtering."""
        rank_metric = config['metric']
        
        # Get BEST STL model as the reference baseline
        stl_df_sorted = stl_df.dropna(subset=[rank_metric])
        if stl_df_sorted.empty:
            return
        
        best_stl = stl_df_sorted.sort_values(by=rank_metric, ascending=False).iloc[0]
        
        # Get ALL MTTL models for this category, sorted by rank metric
        mttl_df_sorted = mttl_df.dropna(subset=[rank_metric])
        if mttl_df_sorted.empty:
            return
        
        mttl_df_sorted = mttl_df_sorted.sort_values(by=rank_metric, ascending=False).reset_index(drop=True)
        
        # Headers: Paradigm | Strategy | Rank | Metric1 | Metric2 | ... | Gain(Metric1) | Gain(Metric2) | ...
        metric_labels = [m.replace('_', ' ').title() for m in config['display_metrics']]
        gain_labels = [f"Gain {label} (%)" for label in metric_labels]
        headers = ['Paradigm', 'Strategy', 'Rank'] + metric_labels + gain_labels
        
        header_row = self.current_row
        self.add_header_row(headers)
        
        # STL row (reference baseline) - no gains for STL
        stl_metric_values = [best_stl.get(m, np.nan) for m in config['display_metrics']]
        stl_gain_values = ['-'] * len(config['display_metrics'])  # No gains for baseline
        stl_values = ['STL', best_stl['strategy'], '★ Best'] + stl_metric_values + stl_gain_values
        self.add_data_row(stl_values)
        
        # MTTL rows (all models with individual metric gains vs best STL)
        for idx, (_, mttl_row) in enumerate(mttl_df_sorted.iterrows(), 1):
            # Metric values for this MTTL model
            mttl_metric_values = [mttl_row.get(m, np.nan) for m in config['display_metrics']]
            
            # Calculate INDIVIDUAL gains for each metric
            mttl_gain_values = []
            for metric in config['display_metrics']:
                stl_val = best_stl.get(metric, np.nan)
                mttl_val = mttl_row.get(metric, np.nan)
                
                if pd.notna(stl_val) and pd.notna(mttl_val) and stl_val != 0:
                    gain = ((mttl_val - stl_val) / stl_val) * 100
                    mttl_gain_values.append(gain)
                else:
                    mttl_gain_values.append(np.nan)
            
            mttl_values = ['MTTL', mttl_row['strategy'], f'#{idx}'] + mttl_metric_values + mttl_gain_values
            self.add_data_row(mttl_values)
        
        # Apply color coding to gain columns for MTTL rows
        num_mttl_rows = len(mttl_df_sorted)
        for mttl_row_offset in range(num_mttl_rows):
            actual_row = header_row + 2 + mttl_row_offset  # +1 for header, +1 for STL baseline
            
            for metric_idx, metric in enumerate(config['display_metrics']):
                col_idx = 4 + len(config['display_metrics']) + metric_idx  # Gain columns start here
                cell = self.ws.cell(row=actual_row, column=col_idx)
                
                if isinstance(cell.value, (int, float)) and not np.isnan(cell.value):
                    gain_value = cell.value
                    
                    # Color based on gain: green for positive, red for negative
                    if gain_value > 0:
                        cell.fill = PatternFill(start_color=self.colors.GAIN_POS, 
                                              end_color=self.colors.GAIN_POS, fill_type='solid')
                    elif gain_value < 0:
                        cell.fill = PatternFill(start_color=self.colors.GAIN_NEG, 
                                              end_color=self.colors.GAIN_NEG, fill_type='solid')
        
        # Gain summary row: Show average gains across all MTTL models for each metric
        gain_summary_row = self.current_row
        self.ws.cell(row=gain_summary_row, column=1).value = 'Avg Gain (%)'
        self.ws.cell(row=gain_summary_row, column=1).font = self.fonts.bold
        self.ws.cell(row=gain_summary_row, column=2).value = 'All MTTL'
        self.ws.cell(row=gain_summary_row, column=3).value = '-'
        
        # Merge metric columns (no values there for summary row)
        for col_idx in range(4, 4 + len(config['display_metrics'])):
            cell = self.ws.cell(row=gain_summary_row, column=col_idx)
            cell.value = '-'
            cell.font = self.fonts.bold
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Calculate and display average gain for EACH metric across all MTTL models
        for metric_idx, metric in enumerate(config['display_metrics']):
            col_idx = 4 + len(config['display_metrics']) + metric_idx  # Gain columns start after metric columns
            cell = self.ws.cell(row=gain_summary_row, column=col_idx)
            
            stl_val = best_stl.get(metric, np.nan)
            
            # Collect gains for this metric across all MTTL models
            metric_gains = []
            for _, mttl_row in mttl_df_sorted.iterrows():
                mttl_val = mttl_row.get(metric, np.nan)
                
                if pd.notna(stl_val) and pd.notna(mttl_val) and stl_val != 0:
                    gain = ((mttl_val - stl_val) / stl_val) * 100
                    metric_gains.append(gain)
            
            if metric_gains:
                avg_metric_gain = np.mean(metric_gains)
                cell.value = round(avg_metric_gain, 2)
                cell.number_format = '0.00'
                
                # Color based on gain
                if avg_metric_gain > 0:
                    cell.fill = PatternFill(start_color=self.colors.GAIN_POS, 
                                        end_color=self.colors.GAIN_POS, fill_type='solid')
                elif avg_metric_gain < 0:
                    cell.fill = PatternFill(start_color=self.colors.GAIN_NEG, 
                                        end_color=self.colors.GAIN_NEG, fill_type='solid')
            else:
                cell.value = '-'
            
            cell.font = self.fonts.bold
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        self.ws.row_dimensions[gain_summary_row].height = 18
        self.current_row += 1
        
        # Add Excel AutoFilter to entire table for sorting capability
        table_start_col = 1
        table_end_col = 3 + len(config['display_metrics']) + len(config['display_metrics'])
        table_end_row = self.current_row - 2
        
        filter_range = f"{get_column_letter(table_start_col)}{header_row}:{get_column_letter(table_end_col)}{table_end_row}"
        
        try:
            self.ws.auto_filter.ref = filter_range
        except:
            pass  # AutoFilter 
             
class MetadataSheetBuilder(ExcelSheetBuilder):
    """Metadata sheet with all experiment details."""
    
    def __init__(self, workbook: openpyxl.Workbook):
        super().__init__(workbook, "Metadata")
    
    def build(self, report_gen) -> None:
        """Build metadata sheet."""
        self.add_title("COMPLETE EXPERIMENT METADATA & DETAILS")
        
        df = report_gen.all_runs_df.copy()
        
        # Convert non-serializable columns
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].apply(
                    lambda x: ', '.join(str(v) for v in x) if isinstance(x, list) else str(x)
                )
        
        # Select key columns
        key_columns = ['strategy', 'mode', 'status', 'timestamp_start', 
                      'num_classes_detection', 'num_classes_classif']
        metric_cols = [col for col in df.columns 
                      if df[col].dtype in ['float64', 'int64']
                      and col not in ['run_id']]
        
        display_cols = [c for c in key_columns if c in df.columns] + sorted(metric_cols)[:15]
        
        self.add_section("All Experiments Data")
        self.add_table_with_filtering(df, display_cols, include_rank=True)
        
        self.auto_adjust_columns()

class ExcelReportExporter:
    """Main exporter orchestrating all sheet builders."""
    
    def __init__(self, output_dir: str = "../OUTPUT/reports"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    def create_workbook(self, report_generator) -> str:
        """Create and save the complete  Excel workbook."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        print("Building Summary...")
        SummarySheetBuilder(wb).build(report_generator)
        
        print("Building Leaderboards...")
        LeaderboardSheetBuilder(wb).build(report_generator)
        
        print("Building STL Details...")
        STLDetailsSheetBuilder(wb).build(report_generator)
        
        print("Building MTTL Details...")
        MTTLDetailsSheetBuilder(wb).build(report_generator)
        
        print("Building Comparison (STL vs MTTL)...")
        ComparisonSheetBuilder(wb).build(report_generator)
        
        print("Building Metadata...")
        MetadataSheetBuilder(wb).build(report_generator)
        
        filepath = os.path.join(self.output_dir, f"Performance_Report_{self.timestamp}.xlsx")
        wb.save(filepath)
        
        print("\n" + "="*80)
        print("Excel Report Generated Successfully")
        print("="*80)
        print(f"File: {filepath}")
        print(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Sheets: Summary | Leaderboards | STL Details | MTTL Details | STL vs MTTL | Metadata")
        print("="*80 + "\n")
        
        return filepath

def generate_and_export_report(base_dir: str = "experiments", output_dir: str = "reports", print_notebook: bool = False) -> Optional[str]:
    print("\n" + "="*80)
    print("PERFORMANCE ANALYSIS REPORT GENERATOR")
    print("="*80)
    
    print("\n[1/2] Loading experiment data...")
    try:
        report_generator = ReportGenerator(base_dir=base_dir)
        
        if report_generator.all_runs_df.empty:
            print("No completed experiments found. Export skipped.")
            return None
        
        if print_notebook:
            report_generator.generate_all_reports()
        
        print(f"Loaded {len(report_generator.all_runs_df)} experiments")
        
    except Exception as e:
        print(f"Error loading experiments: {e}")
        return None
    
    print("\n[2/2] Exporting to  Excel workbook...")
    try:
        exporter = ExcelReportExporter(output_dir=output_dir)
        filepath = exporter.create_workbook(report_generator)
        return filepath
        
    except Exception as e:
        print(f"Error during Excel export: {e}")
        import traceback
        traceback.print_exc()
        return None

